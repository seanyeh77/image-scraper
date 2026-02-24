"""
Image scraper for downloading images from Google Images based on keywords.
Supports parallel downloads, size validation, and Excel inventory tracking.
"""

import base64
import os
import time
from concurrent.futures import ThreadPoolExecutor, wait
from io import BytesIO
from pathlib import Path
from typing import List, Tuple

import openpyxl
import requests
from bs4 import BeautifulSoup
from PIL import Image
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# Constants
DEFAULT_MIN_WIDTH = 150
DEFAULT_MIN_HEIGHT = 150
DEFAULT_SCROLLS = 10
SCROLL_DELAY = 2
PAGE_LOAD_DELAY = 3
REQUEST_TIMEOUT = 10
IMAGES_DIR = "images"
INPUT_DIR = "input"
KEYWORD_FILE = "key_word_list.txt"
MASTER_INVENTORY_FILE = "master_inventory.xlsx"
IMAGE_CLASS = "YQ4gaf"


def scroll_page(driver: webdriver.Chrome, scrolls: int = 5, delay: float = 1) -> None:
    """Scroll the page a specified number of times with delays between scrolls."""
    for _ in range(scrolls):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(delay)


def verify_image_size(image_data: bytes, min_width: int = 150, min_height: int = 150) -> bool:
    """Verify that image dimensions meet minimum requirements."""
    try:
        image = Image.open(BytesIO(image_data))
        width, height = image.size
        return width >= min_width and height >= min_height
    except Exception:
        return False


def initialize_chrome_driver() -> webdriver.Chrome:
    """Initialize and return a headless Chrome WebDriver."""
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-dev-shm-usage")
    
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def fetch_image_links(driver: webdriver.Chrome, keyword: str) -> List[str]:
    """Fetch image links from Google Images for the given keyword."""
    search_url = f"https://www.google.com/search?q={keyword}&tbm=isch&tbs=isz:l"
    driver.get(search_url)
    time.sleep(PAGE_LOAD_DELAY)
    
    scroll_page(driver, scrolls=DEFAULT_SCROLLS, delay=SCROLL_DELAY)
    
    soup = BeautifulSoup(driver.page_source, "lxml")
    results = soup.find_all("img", {"class": IMAGE_CLASS})
    image_links = []
    for result in results:
        link = result.get("src") or result.get("data-src")
        if link and isinstance(link, str):
            image_links.append(link)
    
    return image_links


def download_image(link: str) -> bytes:
    """Download image data from URL or decode from base64."""
    if link.startswith("data:image"):
        _, encoded = link.split(",", 1)
        return base64.b64decode(encoded)
    else:
        response = requests.get(link, timeout=REQUEST_TIMEOUT)
        return response.content


def save_images(keyword: str, image_links: List[str], min_width: int, min_height: int) -> Tuple[Path, int]:
    """Download and save images that meet size requirements."""
    output_dir = Path(IMAGES_DIR) / keyword
    output_dir.mkdir(parents=True, exist_ok=True)
    
    downloaded_count = 0
    for link in image_links:
        try:
            image_data = download_image(link)
            
            if verify_image_size(image_data, min_width, min_height):
                output_path = output_dir / f"{keyword}_{downloaded_count + 1}.jpg"
                output_path.write_bytes(image_data)
                downloaded_count += 1
        except Exception:
            continue
    
    return output_dir, downloaded_count


def update_excel_inventory(file_path: Path, keyword: str, downloaded_count: int, sheet_name: str = "Inventory") -> None:
    """Update or create Excel inventory file with downloaded image information."""
    # Create workbook if it doesn't exist
    if not file_path.exists():
        wb = openpyxl.Workbook()
        sheet = wb.active
        if sheet is not None:
            sheet.title = sheet_name
            sheet["A1"] = "Keyword"
            sheet["B1"] = "Filename"
        wb.save(file_path)
    
    # Load workbook and get existing filenames
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    existing_filenames = set()
    
    if sheet is not None and sheet.max_row is not None:
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row, 2).value
            if cell_value:
                existing_filenames.add(str(cell_value))
    
    # Add new entries
    if sheet is not None:
        for i in range(1, downloaded_count + 1):
            img_name = f"{keyword}_{i}.jpg"
            if img_name not in existing_filenames:
                sheet.append([keyword, img_name])
    
    wb.save(file_path)


def search_and_download(keyword: str, min_width: int = 150, min_height: int = 150) -> None:
    """Search and download images for a given keyword."""
    print(f"Searching for images of '{keyword}'...")
    
    driver = initialize_chrome_driver()
    
    try:
        image_links = fetch_image_links(driver, keyword)
        print(f"Found {len(image_links)} images for '{keyword}'")
        
        output_dir, downloaded_count = save_images(keyword, image_links, min_width, min_height)
        
        # Update individual inventory
        individual_inventory = output_dir / f"{keyword}_inventory.xlsx"
        update_excel_inventory(individual_inventory, keyword, downloaded_count)
        
        # Update master inventory
        master_inventory = Path(IMAGES_DIR) / MASTER_INVENTORY_FILE
        update_excel_inventory(master_inventory, keyword, downloaded_count)
        
        print(f"Successfully downloaded {downloaded_count} images for '{keyword}' (meeting size requirements)")
    
    finally:
        driver.quit()


def load_keywords(file_path: Path) -> List[str]:
    """Load and parse keywords from the input file."""
    if not file_path.exists():
        file_path.parent.mkdir(parents=True, exist_ok=True)
        file_path.write_text("", encoding="utf-8")
        return []
    
    content = file_path.read_text(encoding="utf-8")
    return [keyword.strip() for keyword in content.split(",") if keyword.strip()]


def main() -> None:
    """Main entry point for the image scraper."""
    keyword_file_path = Path(INPUT_DIR) / KEYWORD_FILE
    keywords = load_keywords(keyword_file_path)
    
    if not keywords:
        print(f"No keywords found in {keyword_file_path}")
        print("Please add comma-separated keywords to the file.")
        return
    
    print(f"Processing {len(keywords)} keywords: {', '.join(keywords)}")
    
    # Execute searches in parallel
    with ThreadPoolExecutor(max_workers=os.cpu_count()) as executor:
        futures = [
            executor.submit(search_and_download, keyword, DEFAULT_MIN_WIDTH, DEFAULT_MIN_HEIGHT)
            for keyword in keywords
        ]
        wait(futures)
    
    print("\nAll downloads completed!")


if __name__ == "__main__":
    main()